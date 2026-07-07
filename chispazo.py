import pandas as pd
import numpy as np
from collections import Counter, defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import warnings
from datetime import datetime
warnings.filterwarnings('ignore')

# ============================================
# CONFIGURACIÓN OPTIMIZADA CON VENTANAS REDUCIDAS
# ============================================
VENTANA_COOCURRENCIA = 400      # Óptimo: 400 sorteos con decaimiento
DECAIMIENTO = 0.995             # Ponderación temporal (más peso a sorteos recientes)
LIFT_MINIMO = 1.3               # Asociación significativa (30% más que el azar)
MIN_APARICIONES = 3             # Mínimo de co-ocurrencias ponderadas

VENTANA_SIMILITUDES = 333       # Para evitar similitudes (último año aprox)
MAX_SIMILITUD = 3               # Permitir hasta 3 números iguales (evita 4 o 5)

# Ventanas reducidas para métodos complementarios
VENTANA_GAPS = 500              # Para análisis de atrasos
VENTANA_MONTECARLO = 800        # Para frecuencias históricas
VENTANA_GENETICO = 600          # Para el algoritmo genético

PRIMOS = {2, 3, 5, 7, 11, 13, 17, 19, 23}

# Colores consola
COLORS = {
    'rojo': '\033[91m', 'verde': '\033[92m', 'amarillo': '\033[93m',
    'azul': '\033[94m', 'morado': '\033[95m', 'cian': '\033[96m',
    'blanco': '\033[97m', 'reset': '\033[0m', 'negrita': '\033[1m'
}

# Colores Excel
EXCEL_COLORS = {
    'header_bg': '4472C4', 'header_font': 'FFFFFF',
    'par_bg': 'E2EFDA', 'borde_color': '000000'
}

def printc(txt, color='blanco', negrita=False):
    style = COLORS.get(color, COLORS['blanco'])
    if negrita:
        style += COLORS['negrita']
    print(f"{style}{txt}{COLORS['reset']}")

def print_header(txt):
    printc("="*80, 'cian')
    printc(txt, 'amarillo', negrita=True)
    printc("="*80, 'cian')

def leer_historico(ruta):
    df = pd.read_excel(ruta, sheet_name="Chispazo")
    df['FECHA'] = df['FECHA'].astype(str)
    return df

def obtener_ultimos(df, n):
    return df.head(n).copy()

def analizar_combinaciones_recientes(df):
    """Extrae las combinaciones de los últimos VENTANA_SIMILITUDES sorteos"""
    ultimos = obtener_ultimos(df, VENTANA_SIMILITUDES)
    combinaciones = []
    for _, row in ultimos.iterrows():
        nums = [row['R1'], row['R2'], row['R3'], row['R4'], row['R5']]
        combinaciones.append(nums)
    printc(f"\n🚫 EVITANDO SIMILITUDES - Últimos {len(ultimos)} sorteos", 'cian')
    printc(f"   ❌ Se evitarán combinaciones con {MAX_SIMILITUD+1} o más números en común", 'amarillo')
    printc(f"   ✅ Se permiten máximo {MAX_SIMILITUD} números iguales", 'verde')
    return combinaciones

def tiene_muchos_comunes(combo, combinaciones_recientes):
    """Retorna True si el combo tiene 4+ números iguales a alguna combinación reciente"""
    conjunto = set(combo)
    for combo_reciente in combinaciones_recientes:
        comunes = len(conjunto.intersection(set(combo_reciente)))
        if comunes >= (MAX_SIMILITUD + 1):
            return True
    return False

def estadisticas_basicas(df):
    sums, pares, bajos = [], [], []
    for _, row in df.iterrows():
        nums = [row['R1'], row['R2'], row['R3'], row['R4'], row['R5']]
        sums.append(sum(nums))
        pares.append(sum(1 for x in nums if x % 2 == 0))
        bajos.append(sum(1 for x in nums if x <= 14))
    return {'suma_media': np.mean(sums), 'suma_std': np.std(sums), 'pares_media': np.mean(pares), 'bajos_media': np.mean(bajos)}

# ============================================
# ANÁLISIS REAL DEL HISTORIAL (AUTO-AJUSTE)
# ============================================
def analizar_distribucion_real(df):
    distribuciones_tercios = Counter()
    primos_por_sorteo = []
    sumas_historicas = []
    
    for _, row in df.iterrows():
        nums = [row['R1'], row['R2'], row['R3'], row['R4'], row['R5']]
        t1 = sum(1 for x in nums if x <= 9)
        t2 = sum(1 for x in nums if 10 <= x <= 18)
        t3 = sum(1 for x in nums if x >= 19)
        distribuciones_tercios[(t1, t2, t3)] += 1
        primos = sum(1 for x in nums if x in PRIMOS)
        primos_por_sorteo.append(primos)
        sumas_historicas.append(sum(nums))
    
    total = len(df)
    print_header("📊 ANÁLISIS REAL DE TU HISTORIAL")
    
    printc(f"\n📐 DISTRIBUCIÓN REAL DE TERCIOS (sobre {total} sorteos):", 'amarillo')
    distribuciones_aceptadas = []
    for (t1, t2, t3), count in sorted(distribuciones_tercios.items(), key=lambda x: x[1], reverse=True)[:8]:
        pct = count/total*100
        printc(f"   {t1}-{t2}-{t3}: {count} veces ({pct:.1f}%)", 'blanco')
        if pct >= 5:
            distribuciones_aceptadas.append((t1, t2, t3))
    
    printc(f"\n🔢 DISTRIBUCIÓN REAL DE NÚMEROS PRIMOS:", 'amarillo')
    primos_dist = Counter(primos_por_sorteo)
    primos_aceptados = []
    for num, count in sorted(primos_dist.items()):
        pct = count/total*100
        printc(f"   {num} primos: {count} veces ({pct:.1f}%)", 'blanco')
        if pct >= 5:
            primos_aceptados.append(num)
    
    percentiles = np.percentile(sumas_historicas, [10, 25, 50, 75, 90, 95])
    printc(f"\n🎯 PERCENTILES REALES DE SUMAS (basados en TU historial):", 'amarillo')
    printc(f"   Percentil 10: {int(percentiles[0])}", 'blanco')
    printc(f"   Percentil 25: {int(percentiles[1])}", 'blanco')
    printc(f"   Percentil 50: {int(percentiles[2])} (MEDIANA)", 'verde', negrita=True)
    printc(f"   Percentil 75: {int(percentiles[3])}", 'blanco')
    printc(f"   Percentil 90: {int(percentiles[4])}", 'blanco')
    printc(f"   Percentil 95: {int(percentiles[5])}", 'blanco')
    
    return distribuciones_aceptadas, primos_aceptados, [int(p) for p in percentiles]

def validar_combo_real(combo, distribuciones_aceptadas, primos_aceptados):
    t1 = sum(1 for x in combo if x <= 9)
    t2 = sum(1 for x in combo if 10 <= x <= 18)
    t3 = sum(1 for x in combo if x >= 19)
    if (t1, t2, t3) not in distribuciones_aceptadas:
        return False
    primos = sum(1 for x in combo if x in PRIMOS)
    if primos not in primos_aceptados:
        return False
    return True

# ============================================
# 1. CO-OCURRENCIA AVANZADA (MÉTODO PRINCIPAL)
# ============================================
def analizar_gaps_simple(df, ventana=VENTANA_GAPS):
    """Retorna gaps SOLO en los últimos N sorteos"""
    df_temp = df.head(ventana).copy()
    df_temp['FECHA_DT'] = pd.to_datetime(df_temp['FECHA'], format='%d/%m/%Y', errors='coerce')
    df_asc = df_temp.sort_values('FECHA_DT', ascending=True).reset_index(drop=True)
    ultima = {num: -1 for num in range(1, 29)}
    for idx, row in df_asc.iterrows():
        for col in ['R1','R2','R3','R4','R5']:
            ultima[row[col]] = idx
    gaps = {}
    for num in range(1, 29):
        if ultima[num] == -1:
            gaps[num] = len(df_asc)
        else:
            gaps[num] = len(df_asc) - ultima[num] - 1
    return gaps

def coocurrencia_avanzada(df, combinaciones_recientes, distribuciones_aceptadas, primos_aceptados, tipo='caliente'):
    """
    tipo: 'caliente' (semilla con mayor frecuencia)
          'frio'    (semilla con mayor gap)
          'tripleta' (basado en tripletas reales)
    """
    df_ventana = df.head(VENTANA_COOCURRENCIA).copy()
    total = len(df_ventana)
    
    # Pesos exponenciales (más reciente = mayor peso)
    pesos = [DECAIMIENTO ** i for i in range(total)]
    pesos = np.array(pesos) / sum(pesos)
    
    # Matriz de co-ocurrencia ponderada
    matriz_cooc = np.zeros((29, 29))
    freqs = np.zeros(29)
    
    for idx, (_, row) in enumerate(df_ventana.iterrows()):
        nums = [row['R1'], row['R2'], row['R3'], row['R4'], row['R5']]
        peso = pesos[idx]
        for i in range(5):
            freqs[nums[i]] += peso
            for j in range(i+1, 5):
                a, b = nums[i], nums[j]
                matriz_cooc[a][b] += peso
                matriz_cooc[b][a] += peso
    
    # Normalizar frecuencias para probabilidades
    prob_ind = freqs / (total * 5)
    
    # Calcular Lift para cada par
    lift = np.zeros((29, 29))
    pares_fuertes = []
    for i in range(1, 29):
        for j in range(i+1, 29):
            p_i = prob_ind[i]
            p_j = prob_ind[j]
            p_ij = matriz_cooc[i][j] / total
            if p_i > 0 and p_j > 0 and p_ij > 0:
                lift_ij = p_ij / (p_i * p_j)
                lift[i][j] = lift_ij
                lift[j][i] = lift_ij
                if matriz_cooc[i][j] >= MIN_APARICIONES and lift_ij >= LIFT_MINIMO:
                    pares_fuertes.append(((i, j), matriz_cooc[i][j], lift_ij))
    
    printc(f"\n🕸️ CO-OCURRENCIA ({tipo}) - Ventana {total} sorteos, decaimiento {DECAIMIENTO}", 'morado', negrita=True)
    printc(f"   Pares fuertes encontrados: {len(pares_fuertes)}", 'blanco')
    for (a,b), count, l in pares_fuertes[:5]:
        printc(f"     {a:2d} ↔ {b:2d} | frecuencia: {count:.1f} | lift: {l:.2f}", 'blanco')
    
    # Selección de semilla según tipo
    if tipo == 'caliente':
        semilla = int(np.argmax(freqs[1:]) + 1)
        printc(f"   Semilla caliente: {semilla} (frecuencia: {freqs[semilla]:.1f})", 'verde')
    elif tipo == 'frio':
        gaps = analizar_gaps_simple(df)
        semilla = max(gaps, key=gaps.get)
        printc(f"   Semilla fría (atrasada): {semilla} (gap: {gaps[semilla]} sorteos)", 'verde')
    else:  # tripleta
        tripletas = defaultdict(float)
        for idx, (_, row) in enumerate(df_ventana.iterrows()):
            nums = sorted([row['R1'], row['R2'], row['R3'], row['R4'], row['R5']])
            peso = pesos[idx]
            for i in range(5):
                for j in range(i+1, 5):
                    for k in range(j+1, 5):
                        tripletas[(nums[i], nums[j], nums[k])] += peso
        if tripletas:
            mejor_tripleta = max(tripletas, key=tripletas.get)
            semilla = mejor_tripleta[0]
            printc(f"   Mejor tripleta: {mejor_tripleta} (frecuencia: {tripletas[mejor_tripleta]:.1f})", 'verde')
        else:
            semilla = np.random.randint(1, 29)
    
    # Construir combinación a partir de la semilla y sus mejores compañeros
    compañeros = []
    for (a,b), count, l in pares_fuertes:
        if a == semilla:
            compañeros.append((b, count, l))
        elif b == semilla:
            compañeros.append((a, count, l))
    compañeros.sort(key=lambda x: x[2], reverse=True)
    
    combo = {semilla}
    for comp, _, _ in compañeros:
        if len(combo) < 5 and comp not in combo:
            combo.add(comp)
        if len(combo) == 5:
            break
    
    # Si faltan números, completar con los más probables individualmente
    if len(combo) < 5:
        restantes = [n for n in range(1, 29) if n not in combo]
        probs_rest = [freqs[n] for n in restantes]
        if sum(probs_rest) > 0:
            probs_rest = np.array(probs_rest) / sum(probs_rest)
            while len(combo) < 5:
                nuevo = np.random.choice(restantes, p=probs_rest)
                combo.add(nuevo)
    
    combo_final = sorted(combo)
    
    # Validar similitudes y reglas
    if tiene_muchos_comunes(combo_final, combinaciones_recientes):
        for _ in range(3):
            combo_final = sorted(combo)
            # Cambiar un número problemático
            for i in range(5):
                nuevo_num = np.random.randint(1, 29)
                while nuevo_num in combo_final:
                    nuevo_num = np.random.randint(1, 29)
                combo_final[i] = nuevo_num
                combo_final.sort()
                if not tiene_muchos_comunes(combo_final, combinaciones_recientes):
                    break
    
    if not validar_combo_real(combo_final, distribuciones_aceptadas, primos_aceptados):
        for _ in range(5):
            combo_final = sorted(np.random.choice(range(1, 29), 5, replace=False))
            if validar_combo_real(combo_final, distribuciones_aceptadas, primos_aceptados) and not tiene_muchos_comunes(combo_final, combinaciones_recientes):
                break
    
    return combo_final

# ============================================
# 2. GAPS (NÚMEROS ATRASADOS) CON VENTANA REDUCIDA
# ============================================
def generar_gaps(df, combinaciones_recientes, distribuciones_aceptadas, primos_aceptados, objetivo_suma):
    gaps = analizar_gaps_simple(df, ventana=VENTANA_GAPS)
    ordenados = sorted(gaps.items(), key=lambda x: x[1], reverse=True)
    printc(f"\n⏳ GAPS - Últimos {VENTANA_GAPS} sorteos", 'azul', negrita=True)
    printc(f"   Top números más atrasados:", 'blanco')
    for num, gap in ordenados[:7]:
        printc(f"     Nº {num:2d}: {gap} sorteos sin salir", 'blanco')
    
    mejor_combo = None
    mejor_distancia = float('inf')
    
    for _ in range(2000):
        sel = set()
        # Añadir algunos de los más atrasados
        for num, _ in ordenados[:5]:
            if np.random.random() < 0.5 and len(sel) < 5:
                sel.add(num)
        # Completar con números aleatorios (priorizando atrasados)
        restantes = [n for n in range(1, 29) if n not in sel]
        while len(sel) < 5:
            if np.random.random() < 0.7 and len(ordenados) > 0:
                num = ordenados[np.random.randint(0, min(10, len(ordenados)))][0]
            else:
                num = np.random.randint(1, 29)
            sel.add(num)
        
        combo = sorted(sel)
        
        if tiene_muchos_comunes(combo, combinaciones_recientes):
            continue
        if not validar_combo_real(combo, distribuciones_aceptadas, primos_aceptados):
            continue
        
        suma = sum(combo)
        distancia = abs(suma - objetivo_suma)
        if distancia < mejor_distancia:
            mejor_distancia = distancia
            mejor_combo = combo
            if distancia == 0:
                break
    
    if mejor_combo is None:
        mejor_combo = sorted(np.random.choice(range(1, 29), 5, replace=False))
    
    return mejor_combo

# ============================================
# 3. MONTE CARLO CON VENTANA REDUCIDA
# ============================================
def generar_montecarlo(df, combinaciones_recientes, distribuciones_aceptadas, primos_aceptados, objetivo_suma):
    printc(f"\n🎲 MONTE CARLO - Últimos {VENTANA_MONTECARLO} sorteos", 'verde', negrita=True)
    df_ventana = df.head(VENTANA_MONTECARLO)
    
    todas_frecuencias = Counter()
    for _, row in df_ventana.iterrows():
        for col in ['R1','R2','R3','R4','R5']:
            todas_frecuencias[row[col]] += 1
    
    top_reales = sorted(todas_frecuencias.items(), key=lambda x: x[1], reverse=True)[:10]
    printc(f"   📊 Números más frecuentes: {[num for num, _ in top_reales]}", 'blanco')
    
    pesos = [todas_frecuencias.get(num, 0) + 1 for num in range(1, 29)]
    pesos = np.array(pesos) / sum(pesos)
    nums = list(range(1, 29))
    
    mejor_combo = None
    mejor_distancia = float('inf')
    
    for _ in range(3000):
        sel = set()
        while len(sel) < 5:
            sel.add(np.random.choice(nums, p=pesos))
        combo = sorted(sel)
        
        if tiene_muchos_comunes(combo, combinaciones_recientes):
            continue
        if not validar_combo_real(combo, distribuciones_aceptadas, primos_aceptados):
            continue
        
        suma = sum(combo)
        distancia = abs(suma - objetivo_suma)
        if distancia < mejor_distancia:
            mejor_distancia = distancia
            mejor_combo = combo
    
    if mejor_combo is None:
        mejor_combo = sorted(np.random.choice(range(1, 29), 5, replace=False))
    
    printc(f"   ✅ Combinación generada: {mejor_combo} (suma: {sum(mejor_combo)})", 'verde')
    return mejor_combo

# ============================================
# 4. ALGORITMO GENÉTICO CON VENTANA REDUCIDA
# ============================================
def generar_poblacion_inicial(tamano, distribuciones_aceptadas, primos_aceptados, stats, combinaciones_recientes, df_ventana):
    poblacion = []
    for _ in range(tamano):
        for __ in range(500):
            combo = sorted(np.random.choice(range(1, 29), 5, replace=False))
            if tiene_muchos_comunes(combo, combinaciones_recientes):
                continue
            if validar_combo_real(combo, distribuciones_aceptadas, primos_aceptados):
                suma = sum(combo)
                if abs(suma - stats['suma_media']) <= stats['suma_std'] * 1.5:
                    poblacion.append(combo)
                    break
    return poblacion

def fitness_genetico(combo, df_ventana, stats, distribuciones_aceptadas, primos_aceptados, combinaciones_historicas, combinaciones_recientes):
    if not validar_combo_real(combo, distribuciones_aceptadas, primos_aceptados):
        return 0
    if tiene_muchos_comunes(combo, combinaciones_recientes):
        return 0
    
    fitness = 0
    suma = sum(combo)
    dif_suma = abs(suma - stats['suma_media'])
    fitness += (1 - dif_suma / stats['suma_std'] / 2) * 20
    
    t1 = sum(1 for x in combo if x <= 9)
    t2 = sum(1 for x in combo if 10 <= x <= 18)
    t3 = sum(1 for x in combo if x >= 19)
    if (t1, t2, t3) in [(1,2,2), (2,1,2), (2,2,1)]:
        fitness += 30
    elif (t1, t2, t3) in [(1,1,3), (3,1,1), (1,3,1)]:
        fitness += 15
    
    primos = sum(1 for x in combo if x in PRIMOS)
    if 1 <= primos <= 3:
        fitness += 20
    
    combo_tuple = tuple(combo)
    if combo_tuple in combinaciones_historicas:
        fitness -= 50
    
    return max(0, fitness)

def algoritmo_genetico(df, stats, distribuciones_aceptadas, primos_aceptados, combinaciones_recientes, objetivo_suma, generaciones=40, poblacion_tamano=20):
    printc(f"\n🧬 ALGORITMO GENÉTICO - Últimos {VENTANA_GENETICO} sorteos", 'morado', negrita=True)
    df_ventana = df.head(VENTANA_GENETICO)
    
    combinaciones_historicas = set()
    for _, row in df_ventana.iterrows():
        combo = (row['R1'], row['R2'], row['R3'], row['R4'], row['R5'])
        combinaciones_historicas.add(combo)
    printc(f"   📊 {len(combinaciones_historicas)} combinaciones históricas (últimos {VENTANA_GENETICO} sorteos)", 'blanco')
    
    poblacion = generar_poblacion_inicial(poblacion_tamano, distribuciones_aceptadas, primos_aceptados, stats, combinaciones_recientes, df_ventana)
    
    if len(poblacion) == 0:
        printc("   ⚠️ No se pudo generar población inicial, usando combinación aleatoria", 'amarillo')
        return sorted(np.random.choice(range(1, 29), 5, replace=False))
    
    def fitness_local(combo):
        return fitness_genetico(combo, df_ventana, stats, distribuciones_aceptadas, primos_aceptados, combinaciones_historicas, combinaciones_recientes)
    
    mejor_fitness_historico = 0
    mejor_combo_historico = None
    
    for gen in range(generaciones):
        fitnesses = [fitness_local(combo) for combo in poblacion]
        max_fitness = max(fitnesses)
        if max_fitness > mejor_fitness_historico:
            mejor_fitness_historico = max_fitness
            mejor_combo_historico = poblacion[fitnesses.index(max_fitness)].copy()
        
        poblacion_sorted = [combo for _, combo in sorted(zip(fitnesses, poblacion), key=lambda x: x[0], reverse=True)]
        poblacion = poblacion_sorted[:poblacion_tamano//2]
        
        while len(poblacion) < poblacion_tamano:
            idx1 = np.random.randint(0, min(8, len(poblacion_sorted)))
            idx2 = np.random.randint(0, min(8, len(poblacion_sorted)))
            padre1 = poblacion_sorted[idx1]
            padre2 = poblacion_sorted[idx2]
            
            union = list(set(padre1 + padre2))
            if len(union) >= 5:
                hijo = sorted(np.random.choice(union, 5, replace=False))
            else:
                hijo = padre1.copy()
            
            if np.random.random() < 0.2:
                idx = np.random.randint(0, 5)
                nuevo_num = np.random.randint(1, 29)
                while nuevo_num in hijo:
                    nuevo_num = np.random.randint(1, 29)
                hijo[idx] = nuevo_num
                hijo.sort()
            
            if not tiene_muchos_comunes(hijo, combinaciones_recientes):
                if validar_combo_real(hijo, distribuciones_aceptadas, primos_aceptados):
                    poblacion.append(hijo)
        
        if (gen + 1) % 10 == 0:
            printc(f"   Generación {gen+1}/{generaciones} - Mejor fitness: {mejor_fitness_historico:.1f}", 'cian')
    
    if mejor_combo_historico is None:
        mejor_combo_historico = sorted(np.random.choice(range(1, 29), 5, replace=False))
    
    # Ajuste final por objetivo de suma
    suma_mejor = sum(mejor_combo_historico)
    if abs(suma_mejor - objetivo_suma) > stats['suma_std']:
        for i in range(5):
            for nuevo in range(1, 29):
                if nuevo not in mejor_combo_historico:
                    nuevo_combo = mejor_combo_historico.copy()
                    nuevo_combo[i] = nuevo
                    nuevo_combo.sort()
                    if validar_combo_real(nuevo_combo, distribuciones_aceptadas, primos_aceptados) and not tiene_muchos_comunes(nuevo_combo, combinaciones_recientes):
                        if abs(sum(nuevo_combo) - objetivo_suma) < abs(suma_mejor - objetivo_suma):
                            mejor_combo_historico = nuevo_combo
                            break
                break
    
    printc(f"\n   ✅ Mejor combinación encontrada: {mejor_combo_historico}", 'verde')
    return mejor_combo_historico

# ============================================
# OPTIMIZACIÓN DE COBERTURA FINAL
# ============================================
def optimizar_cobertura_final(combinaciones, stats, distribuciones_aceptadas, primos_aceptados, combinaciones_recientes, max_intentos=80):
    printc(f"\n🎯 OPTIMIZANDO COBERTURA - Buscando 28/28 números", 'cian', negrita=True)
    
    combinaciones = [list(c) for c in combinaciones]
    
    for intento in range(max_intentos):
        numeros_cubiertos = set()
        for c in combinaciones:
            numeros_cubiertos.update(c)
        
        faltantes = [n for n in range(1, 29) if n not in numeros_cubiertos]
        
        if len(faltantes) == 0:
            printc(f"   ✅ ¡COBERTURA PERFECTA! 28/28 números", 'verde', negrita=True)
            break
        
        contador = Counter()
        for c in combinaciones:
            contador.update(c)
        
        repetidos = [num for num, count in contador.items() if count >= 2]
        repetidos.sort(key=lambda x: contador[x], reverse=True)
        
        if not repetidos or not faltantes:
            break
        
        reemplazo_hecho = False
        for num_repetido in repetidos[:5]:
            for num_faltante in faltantes[:5]:
                for i, combo in enumerate(combinaciones):
                    if num_repetido in combo:
                        nuevo_combo = [num_faltante if x == num_repetido else x for x in combo]
                        nuevo_combo.sort()
                        
                        if tiene_muchos_comunes(nuevo_combo, combinaciones_recientes):
                            continue
                        if not validar_combo_real(nuevo_combo, distribuciones_aceptadas, primos_aceptados):
                            continue
                        
                        combinaciones[i] = nuevo_combo
                        reemplazo_hecho = True
                        break
                if reemplazo_hecho:
                    break
            if reemplazo_hecho:
                break
        
        if not reemplazo_hecho:
            break
    
    numeros_cubiertos_final = set()
    for c in combinaciones:
        numeros_cubiertos_final.update(c)
    
    faltantes_final = [n for n in range(1, 29) if n not in numeros_cubiertos_final]
    
    printc(f"\n   📊 RESULTADO FINAL:", 'amarillo', negrita=True)
    printc(f"   Cobertura: {len(numeros_cubiertos_final)}/28 números", 'verde')
    if faltantes_final:
        printc(f"   ❌ Números NO cubiertos: {faltantes_final}", 'rojo')
    else:
        printc(f"   ✅ ¡PERFECTO! Cubriste los 28 números", 'verde', negrita=True)
    
    return [tuple(sorted(c)) for c in combinaciones]

# ============================================
# GUARDADO EN EXCEL CON FORMATO PROFESIONAL
# ============================================
def guardar_output_excel(ruta, combinaciones, stats, metodos, objetivos):
    datos = []
    for i, c in enumerate(combinaciones):
        suma = sum(c)
        dif = suma - stats['suma_media']
        primos_count = sum(1 for x in c if x in PRIMOS)
        t1 = sum(1 for x in c if x <= 9)
        t2 = sum(1 for x in c if 10 <= x <= 18)
        t3 = sum(1 for x in c if x >= 19)
        
        if abs(dif) <= stats['suma_std']:
            status = "✅ EN RANGO"
        elif dif < 0:
            status = "📉 BAJA"
        else:
            status = "📈 ALTA"
        
        datos.append({
            'Combo': i+1,
            'Método': metodos[i],
            'N1': c[0], 'N2': c[1], 'N3': c[2], 'N4': c[3], 'N5': c[4],
            'Suma': suma,
            'Objetivo': objetivos[i] if i < len(objetivos) else "-",
            'Diferencia': f"{dif:+.0f}",
            'Status': status,
            'Pares': sum(1 for x in c if x % 2 == 0),
            'Primos': primos_count,
            '1-9': t1,
            '10-18': t2,
            '19-28': t3
        })
    
    df_nuevo = pd.DataFrame(datos)
    
    with pd.ExcelWriter(ruta, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_nuevo.to_excel(writer, sheet_name='OUTPUT', index=False)
        
        workbook = writer.book
        sheet = writer.sheets['OUTPUT']
        
        header_font = Font(bold=True, size=11, color=EXCEL_COLORS['header_font'])
        header_fill = PatternFill(start_color=EXCEL_COLORS['header_bg'], end_color=EXCEL_COLORS['header_bg'], fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color=EXCEL_COLORS['borde_color']),
            right=Side(style='thin', color=EXCEL_COLORS['borde_color']),
            top=Side(style='thin', color=EXCEL_COLORS['borde_color']),
            bottom=Side(style='thin', color=EXCEL_COLORS['borde_color'])
        )
        
        column_widths = {
            'A': 8, 'B': 22, 'C': 6, 'D': 6, 'E': 6, 'F': 6, 'G': 6,
            'H': 8, 'I': 10, 'J': 12, 'K': 14, 'L': 8, 'M': 8, 'N': 6, 'O': 8, 'P': 8
        }
        
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
        
        for col in range(1, len(df_nuevo.columns) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row in range(2, len(df_nuevo) + 2):
            for col in range(1, len(df_nuevo.columns) + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color=EXCEL_COLORS['par_bg'], end_color=EXCEL_COLORS['par_bg'], fill_type='solid')
                
                if col == 11:
                    status_value = cell.value
                    if status_value == "✅ EN RANGO":
                        cell.font = Font(color='006100', bold=True)
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif status_value == "📉 BAJA":
                        cell.font = Font(color='9C6500', bold=True)
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    elif status_value == "📈 ALTA":
                        cell.font = Font(color='9C5700', bold=True)
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                
                if 3 <= col <= 7:
                    cell.font = Font(bold=True, size=11)
    
    printc(f"\n✅ OUTPUT guardado en {ruta} con formato profesional", 'verde', negrita=True)

def guardar_repetidas(ruta, df):
    df_original = pd.read_excel(ruta, sheet_name="Chispazo")
    df_original['FECHA'] = df_original['FECHA'].astype(str)
    grupos = defaultdict(list)
    for idx, row in df_original.iterrows():
        nums = [row['R1'], row['R2'], row['R3'], row['R4'], row['R5']]
        combo = tuple(nums)
        grupos[combo].append({'fecha': row['FECHA'], 'concurso': row['CONCURSO'], 'idx_original': idx})
    repetidas = {c: a for c, a in grupos.items() if len(a) > 1}
    if not repetidas:
        printc("✅ No hay combinaciones repetidas", 'verde')
        return
    datos = []
    for combo, apps in repetidas.items():
        apps_ordenadas = sorted(apps, key=lambda x: x['idx_original'], reverse=False)
        veces = len(apps_ordenadas)
        fila = {'N1': combo[0], 'N2': combo[1], 'N3': combo[2], 'N4': combo[3], 'N5': combo[4],
                'Combinacion': f"{combo[0]:02d}-{combo[1]:02d}-{combo[2]:02d}-{combo[3]:02d}-{combo[4]:02d}",
                'Veces': veces}
        for i, app in enumerate(apps_ordenadas, 1):
            fila[f'{i}° Fecha'] = app['fecha']
            fila[f'{i}° Concurso'] = int(app['concurso'])
        try:
            f1 = datetime.strptime(apps_ordenadas[0]['fecha'], '%d/%m/%Y')
            f2 = datetime.strptime(apps_ordenadas[-1]['fecha'], '%d/%m/%Y')
            fila['Dias_Total'] = abs((f2 - f1).days)
        except:
            fila['Dias_Total'] = 0
        datos.append(fila)
    df_repes = pd.DataFrame(datos)
    fechas_recientes = []
    for _, row in df_repes.iterrows():
        ultima_fecha = row[f"{int(row['Veces'])}° Fecha"]
        try:
            d, m, a = ultima_fecha.split('/')
            fechas_recientes.append(int(f"{a}{m}{d}"))
        except:
            fechas_recientes.append(0)
    df_repes['_fecha_ordenar'] = fechas_recientes
    df_repes = df_repes.sort_values(['Veces', '_fecha_ordenar'], ascending=[False, False])
    df_repes = df_repes.drop(columns=['_fecha_ordenar'])
    
    with pd.ExcelWriter(ruta, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_repes.to_excel(writer, sheet_name='COMBINACIONES_REPETIDAS', index=False)
        
        workbook = writer.book
        sheet = writer.sheets['COMBINACIONES_REPETIDAS']
        
        header_font = Font(bold=True, size=11, color=EXCEL_COLORS['header_font'])
        header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col in range(1, len(df_repes.columns) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        for row in range(2, len(df_repes) + 2):
            for col in range(1, len(df_repes.columns) + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            sheet.column_dimensions[col_letter].width = adjusted_width
    
    printc(f"\n✅ COMBINACIONES_REPETIDAS guardada ({len(df_repes)} combinaciones)", 'verde')

# ============================================
# VISUALIZACIÓN EN CONSOLA
# ============================================
def mostrar_resultados_final(combinaciones, stats, metodos, objetivos, distribuciones_aceptadas, primos_aceptados):
    print_header("🎲 CHISPAZO VERSIÓN 9.0 - CO-OCURRENCIA TOTAL 🕸️")
    
    printc(f"\n📊 ESTADÍSTICAS BASE:", 'cian', negrita=True)
    printc(f"   Suma promedio: {stats['suma_media']:.1f} (±{stats['suma_std']:.1f})", 'blanco')
    printc(f"   Pares promedio: {stats['pares_media']:.1f}", 'blanco')
    printc(f"   Bajos (1-14) promedio: {stats['bajos_media']:.1f}", 'blanco')
    
    printc(f"\n📐 FILTROS AUTO-AJUSTADOS:", 'verde', negrita=True)
    printc(f"   Distribuciones de tercios aceptadas: {distribuciones_aceptadas}", 'blanco')
    printc(f"   Cantidades de primos aceptadas: {primos_aceptados}", 'blanco')
    printc(f"   🚫 EVITA combinaciones con 4 o más números iguales a últimos {VENTANA_SIMILITUDES} sorteos", 'amarillo')
    
    numeros_cubiertos = set()
    for c in combinaciones:
        numeros_cubiertos.update(c)
    cobertura = len(numeros_cubiertos)
    
    printc(f"\n🎯 COBERTURA TOTAL: {cobertura}/28 números ({cobertura/28*100:.1f}%)", 'verde', negrita=True)
    faltantes = [n for n in range(1, 29) if n not in numeros_cubiertos]
    if faltantes:
        printc(f"   ❌ Números NO cubiertos ({len(faltantes)}): {faltantes}", 'rojo')
    else:
        printc(f"   ✅ ¡PERFECTO! Cubriste los 28 números", 'verde', negrita=True)
    
    printc(f"\n🎯 MIS 6 COMBINACIONES GANADORAS:", 'amarillo', negrita=True)
    printc("┌────────────────────────────────────────────────────────────────────┐", 'cian')
    
    for i, c in enumerate(combinaciones):
        suma = sum(c)
        dif = suma - stats['suma_media']
        primos_count = sum(1 for x in c if x in PRIMOS)
        t1 = sum(1 for x in c if x <= 9)
        t2 = sum(1 for x in c if 10 <= x <= 18)
        t3 = sum(1 for x in c if x >= 19)
        
        printc(f"│ {metodos[i]}", 'morado', negrita=True)
        printc(f"│   🎲 {c[0]:2d} - {c[1]:2d} - {c[2]:2d} - {c[3]:2d} - {c[4]:2d}", 'blanco', negrita=True)
        printc(f"│   Suma: {suma:3d} | Pares: {sum(1 for x in c if x%2==0)} | Primos: {primos_count}", 'blanco')
        printc(f"│   Tercios: {t1}-{t2}-{t3}", 'blanco')
        if i < len(combinaciones) - 1:
            printc("├────────────────────────────────────────────────────────────────────┤", 'cian')
    
    printc("└────────────────────────────────────────────────────────────────────┘", 'cian')
    
    print_header("✅ ¡SUERTE CON ESTAS 6 COMBINACIONES! 🍀")
    printc("📌 Basadas en: 3 Co-ocurrencias (caliente, fría, tripleta) + Gaps + Monte Carlo + Genético", 'verde')
    printc("📌 Co-ocurrencia avanzada: ventana 400 sorteos, decaimiento 0.995, lift mínimo 1.3", 'cian')
    printc(f"📌 EVITA combinaciones con 4+ números iguales a últimos {VENTANA_SIMILITUDES} sorteos", 'amarillo')
    printc("📌 Optimizadas para máxima cobertura de números", 'verde')
    printc("📌 Formato profesional en Excel con colores y bordes", 'cian')
    printc("\n🐐 Hecho por la CABRA de las matemáticas 🐐", 'amarillo', negrita=True)

# ============================================
# MAIN
# ============================================
def main():
    archivo = "chispazo_historico.xlsx"
    try:
        printc("\n🚀 CHISPAZO VERSIÓN 9.0 - CO-OCURRENCIA TOTAL 🕸️", 'amarillo', negrita=True)
        printc("   🔥 3 métodos de CO-OCURRENCIA avanzada (caliente, fría, tripleta)", 'cian')
        printc("   📐 Ventana de 400 sorteos con decaimiento exponencial (0.995)", 'cian')
        printc("   🎯 Lift mínimo 1.3 - asociaciones REALES entre números", 'cian')
        printc("   📊 + Gaps (500) + Monte Carlo (800) + Genético (600) con ventanas reducidas", 'cian')
        printc("   🚫 EVITA combinaciones con 4 o más números iguales en últimos 333 sorteos", 'rojo')
        printc("   🎯 Buscando cobertura PERFECTA de 28/28 números", 'verde', negrita=True)
        printc("   🎨 Excel con formato profesional", 'cian')
        
        df = leer_historico(archivo)
        combinaciones_recientes = analizar_combinaciones_recientes(df)
        distribuciones_aceptadas, primos_aceptados, percentiles = analizar_distribucion_real(df)
        stats = estadisticas_basicas(df)
        
        # 6 objetivos basados en percentiles reales
        objetivos = percentiles[:6]
        
        printc(f"\n📊 OBJETIVOS DE SUMA ASIGNADOS:", 'amarillo', negrita=True)
        printc(f"   1. Co-oc Caliente: {objetivos[0]} (percentil 10)", 'blanco')
        printc(f"   2. Co-oc Fría:    {objetivos[1]} (percentil 25)", 'blanco')
        printc(f"   3. Co-oc Tripleta: {objetivos[2]} (percentil 50)", 'verde', negrita=True)
        printc(f"   4. Gaps:           {objetivos[3]} (percentil 75)", 'blanco')
        printc(f"   5. Monte Carlo:    {objetivos[4]} (percentil 90)", 'blanco')
        printc(f"   6. Genético:       {objetivos[5]} (percentil 95)", 'blanco')
        
        printc("\n🎲 GENERANDO LAS 6 COMBINACIONES DEFINITIVAS...", 'verde', negrita=True)
        
        # Generar las 6 combinaciones
        c1 = coocurrencia_avanzada(df, combinaciones_recientes, distribuciones_aceptadas, primos_aceptados, tipo='caliente')
        c2 = coocurrencia_avanzada(df, combinaciones_recientes, distribuciones_aceptadas, primos_aceptados, tipo='frio')
        c3 = coocurrencia_avanzada(df, combinaciones_recientes, distribuciones_aceptadas, primos_aceptados, tipo='tripleta')
        c4 = generar_gaps(df, combinaciones_recientes, distribuciones_aceptadas, primos_aceptados, objetivos[3])
        c5 = generar_montecarlo(df, combinaciones_recientes, distribuciones_aceptadas, primos_aceptados, objetivos[4])
        c6 = algoritmo_genetico(df, stats, distribuciones_aceptadas, primos_aceptados, combinaciones_recientes, objetivos[5])
        
        combinaciones = [c1, c2, c3, c4, c5, c6]
        metodos = ['🕸️ Co-oc Caliente', '🕸️ Co-oc Fría', '🕸️ Co-oc Tripleta', '⏳ Gaps', '🎲 Monte Carlo', '🧬 Genético']
        
        # Optimizar cobertura para buscar 28/28
        combinaciones_final = optimizar_cobertura_final(combinaciones, stats, distribuciones_aceptadas, primos_aceptados, combinaciones_recientes, max_intentos=80)
        
        # Guardar resultados
        guardar_output_excel(archivo, combinaciones_final, stats, metodos, objetivos)
        guardar_repetidas(archivo, df)
        mostrar_resultados_final(combinaciones_final, stats, metodos, objetivos, distribuciones_aceptadas, primos_aceptados)
        
    except FileNotFoundError:
        printc(f"\n❌ ERROR: No se encontró '{archivo}'", 'rojo', negrita=True)
        printc("   Asegúrate de que el archivo existe en la misma carpeta", 'amarillo')
    except Exception as e:
        printc(f"\n❌ ERROR: {e}", 'rojo', negrita=True)
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()